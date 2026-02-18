"""
Convert '5-CSP物料_副本.xlsx' into a CSV test data file
that matches the SPM Simulator's default import mapping.

Source: 1041 rows of real CSP production material data.
Output: UTF-8-BOM CSV with Chinese headers matching the app's default mapping.
"""

import csv
import os
import random
import hashlib
from datetime import datetime, timedelta

import openpyxl

random.seed(42)

# ─── Mapping tables ───

# Excel column W (合同性质代码) → app 合同性质
CONTRACT_NATURE_MAP = {
    "QZA": "正式合同",
    "QPA": "期货合同",
    "QSA": "框架协议",
    "YYA": "临时订单",
    "YWA": "临时订单",
    "YZA": "临时订单",
}

# Excel column W → app 合同属性
CONTRACT_ATTR_MAP = {
    "QZA": "现货合同",
    "QPA": "期货合同",
    "QSA": "现货合同",
    "YYA": "过渡材合同",
    "YWA": "其他",
    "YZA": "其他",
}

# Excel column CS (产品大类代码) → app 产品大类
PRODUCT_TYPE_MAP = {
    "I": "结构钢板",
    "X": "热轧板卷",
    "S": "热轧带钢",
    "O": "热轧酸洗板",
}

# Steel grade → hardness level
HARD_GRADES = {
    "BR1500HS", "BST700X", "BST750X", "65Mn", "QStE500TM",
    "B610L", "65W600", "B65A1300",
}
MEDIUM_GRADES = {
    "Q345NQR2", "Q355B", "S355MC", "SAPH440", "SAPH400",
    "SPA-H", "RCL380", "QStE340TM", "S235JR+AR", "SM490A",
    "B530CL", "B50A800",
}
# Everything else → soft

# Steel grade → surface level heuristic
SURFACE_FA_GRADES = {"BR1500HS", "TDC51D+AM C5", "TS350GD+AM C5"}
SURFACE_FB_GRADES = {
    "SAPH440", "SAPH400", "SPHC", "SPHC-MJ", "SPHT1",
    "SAE1006", "HTC2", "BGMY", "RCL380",
}
# Default: FC for structural, FD for special


def get_hardness(grade: str) -> str:
    if grade in HARD_GRADES:
        return "硬"
    if grade in MEDIUM_GRADES:
        return "中"
    return "软"


def get_surface(grade: str, product_code: str) -> str:
    if grade in SURFACE_FA_GRADES:
        return "FA"
    if grade in SURFACE_FB_GRADES:
        return "FB"
    if product_code == "I":
        return "FC"
    return random.choice(["FB", "FC"])


def parse_datetime_14(s: str) -> datetime:
    """Parse 'YYYYMMDDHHmmss' format."""
    s = str(s).strip()
    if len(s) >= 14:
        return datetime.strptime(s[:14], "%Y%m%d%H%M%S")
    raise ValueError(f"Cannot parse datetime: {s}")


def parse_date_8(s: str) -> str:
    """Parse 'YYYYMMDD' format → 'YYYY-MM-DD'."""
    s = str(s).strip()
    if len(s) >= 8:
        d = datetime.strptime(s[:8], "%Y%m%d")
        return d.strftime("%Y-%m-%d")
    return ""


def make_customer_code(name: str) -> str:
    """Generate a stable customer code from name."""
    h = hashlib.md5(name.encode()).hexdigest()
    num = int(h[:4], 16) % 9000 + 1000
    return f"C{num}"


def safe_str(val) -> str:
    """Convert value to stripped string, handle None."""
    if val is None:
        return ""
    return str(val).strip()


def safe_float(val, default=0.0) -> float:
    """Convert value to float, return default if not possible."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0) -> int:
    """Convert value to int, return default if not possible."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, "5-CSP物料_副本.xlsx")
    out_path = os.path.join(script_dir, "csp_materials_1041.csv")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet"]

    # Build column letter → index mapping from header row
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[cell.column_letter] = cell.value

    # We need to shift coiling times to be recent for testing
    # Original data coiling times are from 2022; shift to recent dates
    BASE_DATE = datetime(2026, 2, 17, 0, 0, 0)

    # Collect all original coiling times first to compute offset
    original_times = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        v = {c.column_letter: c.value for c in row}
        ac = safe_str(v.get("AC"))
        if ac and len(ac) >= 14:
            try:
                original_times.append(parse_datetime_14(ac))
            except ValueError:
                pass

    if original_times:
        max_original = max(original_times)
        # Shift so that newest coiling time = BASE_DATE - 1 day
        time_shift = BASE_DATE - timedelta(days=1) - max_original
    else:
        time_shift = timedelta(0)

    fields = [
        "钢卷号", "合同号", "客户名称", "客户代码", "钢种",
        "厚度", "宽度", "重量", "硬度等级", "表面等级",
        "粗糙度要求", "延伸率要求", "产品大类", "合同属性", "合同性质",
        "出口标志", "周交期", "集批代码", "卷取时间", "库龄",
        "库位", "交期", "备注",
    ]

    rows = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        v = {c.column_letter: c.value for c in row}

        # ── Required fields ──
        coil_id = safe_str(v.get("B"))
        if not coil_id:
            skipped += 1
            continue

        steel_grade = safe_str(v.get("M"))
        if not steel_grade:
            skipped += 1
            continue

        thickness = safe_float(v.get("G"))
        width = safe_int(v.get("H"))
        weight = safe_float(v.get("J"))

        if thickness <= 0 or width <= 0 or weight <= 0:
            skipped += 1
            continue

        # Coiling time
        ac_str = safe_str(v.get("AC"))
        if ac_str and len(ac_str) >= 14:
            try:
                coiling_dt = parse_datetime_14(ac_str) + time_shift
                coiling_time = coiling_dt.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                coiling_time = BASE_DATE.strftime("%Y-%m-%d %H:%M:%S")
        else:
            coiling_time = BASE_DATE.strftime("%Y-%m-%d %H:%M:%S")

        # ── Optional fields ──
        contract_no = safe_str(v.get("U"))
        customer_name = safe_str(v.get("BO"))
        if not customer_name:
            customer_name = "未知客户"
        customer_code = make_customer_code(customer_name)

        contract_nature_code = safe_str(v.get("W"))
        contract_nature = CONTRACT_NATURE_MAP.get(contract_nature_code, "正式合同")
        contract_attr = CONTRACT_ATTR_MAP.get(contract_nature_code, "现货合同")

        product_code = safe_str(v.get("CS"))
        product_type = PRODUCT_TYPE_MAP.get(product_code, "热轧板卷")

        export_raw = safe_str(v.get("BL"))
        export_flag = "是" if export_raw == "1" else "否"
        # If contract is export type, also set export flag
        if contract_nature_code == "YWA":
            contract_attr = "出口合同" if export_flag == "是" else "其他"

        weekly_raw = safe_str(v.get("Y"))
        weekly_delivery = "是" if weekly_raw == "Y" else "否"

        batch_code_raw = safe_str(v.get("Z"))
        batch_code = batch_code_raw if batch_code_raw and batch_code_raw != " " else ""
        if batch_code:
            batch_code = f"BTH-{batch_code.zfill(4)}"

        storage_days = safe_int(v.get("AE"), default=0)

        # Storage location: combine 库区 (N) as simplified location
        storage_area = safe_str(v.get("N"))
        if storage_area:
            # Map real storage area codes to app format
            area_map = {
                "H41": "A区-01", "H42": "A区-02", "H43": "A区-03",
                "H44": "A区-04", "H45": "A区-05",
                "H46": "B区-01", "H47": "B区-02",
                "H99": "C区-01", "I91": "D区-01",
            }
            storage_loc = area_map.get(storage_area, f"E区-01")
        else:
            storage_loc = "A区-01"

        # Due date
        due_raw = safe_str(v.get("D"))
        if due_raw and len(due_raw) >= 8:
            try:
                due_dt = datetime.strptime(due_raw[:8], "%Y%m%d")
                # Shift old due dates to be recent
                if due_dt.year < 2025:
                    # Map to future dates relative to BASE_DATE
                    due_dt = BASE_DATE + timedelta(days=random.randint(3, 30))
                due_date = due_dt.strftime("%Y-%m-%d")
            except ValueError:
                due_date = ""
        else:
            due_date = ""

        # Remarks
        remarks_raw = safe_str(v.get("CV"))
        if remarks_raw:
            # Simplify long remarks for the app
            if "脱合同" in remarks_raw:
                remarks = "过渡材"
            elif "余材" in remarks_raw:
                remarks = "库龄较长"
            else:
                remarks = remarks_raw[:20]
        else:
            remarks = ""

        # ── Derived fields ──
        hardness = get_hardness(steel_grade)
        surface = get_surface(steel_grade, product_code)

        roughness_raw = safe_float(v.get("CO"))
        if roughness_raw > 0:
            roughness = f"Ra{roughness_raw}"
        else:
            roughness = ""

        elongation = ""  # Not available in source data

        rows.append({
            "钢卷号": coil_id,
            "合同号": contract_no,
            "客户名称": customer_name,
            "客户代码": customer_code,
            "钢种": steel_grade,
            "厚度": round(thickness, 2),
            "宽度": width,
            "重量": round(weight, 2),
            "硬度等级": hardness,
            "表面等级": surface,
            "粗糙度要求": roughness,
            "延伸率要求": elongation,
            "产品大类": product_type,
            "合同属性": contract_attr,
            "合同性质": contract_nature,
            "出口标志": export_flag,
            "周交期": weekly_delivery,
            "集批代码": batch_code,
            "卷取时间": coiling_time,
            "库龄": storage_days,
            "库位": storage_loc,
            "交期": due_date,
            "备注": remarks,
        })

    # Write CSV
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    # Statistics
    thicknesses = [r["厚度"] for r in rows]
    widths = [r["宽度"] for r in rows]
    weights = [r["重量"] for r in rows]

    print(f"Converted {len(rows)} rows (skipped {skipped}) -> {out_path}")
    print(f"Thickness: min={min(thicknesses)}, max={max(thicknesses)}, avg={sum(thicknesses)/len(thicknesses):.2f}")
    print(f"Width:     min={min(widths)}, max={max(widths)}, avg={sum(widths)/len(widths):.0f}")
    print(f"Weight:    min={min(weights)}, max={max(weights)}, avg={sum(weights)/len(weights):.2f}")

    # Field coverage
    from collections import Counter
    grade_c = Counter(r["钢种"] for r in rows)
    product_c = Counter(r["产品大类"] for r in rows)
    attr_c = Counter(r["合同属性"] for r in rows)
    nature_c = Counter(r["合同性质"] for r in rows)
    hardness_c = Counter(r["硬度等级"] for r in rows)
    surface_c = Counter(r["表面等级"] for r in rows)
    export_c = sum(1 for r in rows if r["出口标志"] == "是")
    weekly_c = sum(1 for r in rows if r["周交期"] == "是")
    batch_c = sum(1 for r in rows if r["集批代码"])
    due_c = sum(1 for r in rows if r["交期"])
    remark_c = sum(1 for r in rows if r["备注"])

    print(f"\n── Coverage Report ──")
    print(f"Steel grades ({len(grade_c)}): {dict(grade_c.most_common(10))} ...")
    print(f"Product types: {dict(product_c)}")
    print(f"Contract attrs: {dict(attr_c)}")
    print(f"Contract natures: {dict(nature_c)}")
    print(f"Hardness: {dict(hardness_c)}")
    print(f"Surface: {dict(surface_c)}")
    print(f"Export: {export_c} ({export_c/len(rows)*100:.1f}%)")
    print(f"Weekly: {weekly_c} ({weekly_c/len(rows)*100:.1f}%)")
    print(f"With batch: {batch_c} ({batch_c/len(rows)*100:.1f}%)")
    print(f"With due date: {due_c} ({due_c/len(rows)*100:.1f}%)")
    print(f"With remarks: {remark_c} ({remark_c/len(rows)*100:.1f}%)")
    print(f"Unique customers: {len(set(r['客户名称'] for r in rows))}")


if __name__ == "__main__":
    main()
